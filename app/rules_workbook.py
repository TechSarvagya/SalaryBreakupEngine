from __future__ import annotations
from typing import Iterable

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import RULES_WORKBOOK_PATH

THIN = Side(border_style="thin", color="9CA3AF")
DOUBLE = Side(border_style="double", color="111827")
TITLE_FILL = PatternFill("solid", fgColor="1F2937")
INPUT_FILL = PatternFill("solid", fgColor="DBEAFE")
OUTPUT_FILL = PatternFill("solid", fgColor="DCFCE7")
RULE_FILL = PatternFill("solid", fgColor="F9FAFB")
GUIDE_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(bold=True, color="FFFFFF")
HEADER_FONT = Font(bold=True, color="111827")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def _autosize_sheet(sheet) -> None:
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 12), 70)
        sheet.column_dimensions[column].width = adjusted_width


def _set_border(cell, *, left=None, right=None, top=None, bottom=None) -> None:
    current = cell.border
    cell.border = Border(
        left=left or current.left,
        right=right or current.right,
        top=top or current.top,
        bottom=bottom or current.bottom,
    )


def _write_instructions_sheet(sheet) -> None:
    rows = [
        ("Salary Rules Workbook", "Use this file to change payroll rules without changing code."),
        ("Safe to edit", "Rule values and formulas inside the colored decision tables. This workbook uses the New tax regime."),
        ("Do not rename", "Sheet names, table names, column headings, or the hidden Glossary structure."),
        ("Input columns", "Blue columns are conditions. Use '-' when the rule should always match."),
        ("Output columns", "Green columns are calculated values. Edit percentages, caps, fixed amounts, and slab formulas here."),
        ("Adding slabs", "Add a new row inside the relevant decision table and keep the same columns and border style."),
        ("After editing", "Upload this workbook in the app and refresh rules before processing employees."),
    ]
    for row_index, (heading, detail) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=heading)
        sheet.cell(row=row_index, column=2, value=detail)
        sheet.cell(row=row_index, column=1).font = HEADER_FONT
        sheet.cell(row=row_index, column=1).fill = GUIDE_FILL
        sheet.cell(row=row_index, column=2).alignment = BODY_ALIGNMENT
    sheet.freeze_panes = "A2"
    _autosize_sheet(sheet)


def _write_table(
    sheet,
    start_row: int,
    table_name: str,
    inputs: Iterable[str],
    outputs: Iterable[str],
    rules: list[list[str]],
    hit_policy: str = "U",
    note: str | None = None,
) -> int:
    inputs = list(inputs)
    outputs = list(outputs)

    title_cell = sheet.cell(row=start_row, column=1, value=table_name)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = BODY_ALIGNMENT
    if note:
        title_cell.comment = Comment(note, "Salary Rule Engine")

    col = 1
    hit_cell = sheet.cell(row=start_row + 1, column=col, value=hit_policy)
    hit_cell.font = HEADER_FONT
    hit_cell.fill = GUIDE_FILL
    hit_cell.comment = Comment("Hit policy used by pyDMNrules. Usually leave this unchanged.", "Salary Rule Engine")
    _set_border(hit_cell, bottom=DOUBLE)
    col += 1
    for inp in inputs:
        header_cell = sheet.cell(row=start_row + 1, column=col, value=inp)
        header_cell.font = HEADER_FONT
        header_cell.fill = INPUT_FILL
        header_cell.alignment = BODY_ALIGNMENT
        header_cell.comment = Comment("Input condition column. '-' means this input does not restrict the rule.", "Salary Rule Engine")
        _set_border(header_cell, bottom=DOUBLE)
        col += 1
    for out in outputs:
        header_cell = sheet.cell(row=start_row + 1, column=col, value=out)
        header_cell.font = HEADER_FONT
        header_cell.fill = OUTPUT_FILL
        header_cell.alignment = BODY_ALIGNMENT
        header_cell.comment = Comment("Output calculation column. This is where the rule returns a value.", "Salary Rule Engine")
        _set_border(header_cell, bottom=DOUBLE)
        col += 1

    if inputs:
        last_input_col = 1 + len(inputs)
        first_output_col = last_input_col + 1
        _set_border(sheet.cell(row=start_row + 1, column=last_input_col), right=DOUBLE)
        _set_border(sheet.cell(row=start_row + 1, column=first_output_col), left=DOUBLE)
    else:
        _set_border(hit_cell, right=DOUBLE)
        _set_border(sheet.cell(row=start_row + 1, column=2), left=DOUBLE)

    last_output_col = 1 + len(inputs) + len(outputs)
    _set_border(sheet.cell(row=start_row + 1, column=last_output_col), right=DOUBLE)

    row = start_row + 2
    for rule in rules:
        for c, val in enumerate(rule, start=1):
            cell = sheet.cell(row=row, column=c, value=val)
            cell.fill = RULE_FILL
            cell.alignment = BODY_ALIGNMENT
            _set_border(cell, left=THIN, right=THIN, top=THIN, bottom=THIN)
            if c == 1:
                cell.comment = Comment("Rule number. Keep unique within this table.", "Salary Rule Engine")
            elif c > 1 + len(inputs):
                cell.comment = Comment("Editable output expression. Change rates, caps, fixed amounts, or slab formulas carefully.", "Salary Rule Engine")
        row += 1

    return row + 1


def _write_glossary(sheet) -> None:
    sheet["A1"] = "Glossary Salary Engine"
    sheet["B1"] = "Business Concept"
    sheet["C1"] = "Attribute"
    sheet.append(["Variable", "Business Concept", "Attribute"])
    rows = [
        ("Employee ID", "Employee", "employee_id"),
        ("Employee Name", None, "employee_name"),
        ("Monthly CTC", None, "monthly_ctc"),
        ("CTC", None, "ctc"),
        ("CCA", None, "cca"),
        ("PF Enabled", None, "pf_enabled"),
        ("State", None, "state"),
        ("Other Deductions", None, "other_deductions"),
        ("Basic", "Salary", "basic"),
        ("HRA", None, "hra"),
        ("Transport Allowance", None, "transport_allowance"),
        ("Medical Allowance", None, "medical_allowance"),
        ("Bonus", None, "bonus"),
        ("Gross Before ESI", None, "gross_before_esi"),
        ("Special", None, "special"),
        ("Gross", None, "gross"),
        ("Employer PF", "EmployerCost", "employer_pf"),
        ("Employer ESI", None, "employer_esi"),
        ("Gratuity", None, "gratuity"),
        ("Employer Insurance", None, "employer_insurance"),
        ("Employee PF", "Deduction", "employee_pf"),
        ("Employee ESI", None, "employee_esi"),
        ("Professional Tax", None, "professional_tax"),
        ("Taxable Annual", "Tax", "taxable_annual"),
        ("Tax Before Rebate", None, "tax_before_rebate"),
        ("Tax After Rebate", None, "tax_after_rebate"),
        ("TDS", None, "tds"),
        ("Take Home", "Final", "take_home"),
        ("CTC (Total)", None, "ctc_total"),
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].fill = TITLE_FILL
    sheet["A2"].font = HEADER_FONT
    for cell in sheet[2]:
        cell.fill = GUIDE_FILL
        cell.font = HEADER_FONT
    sheet.freeze_panes = "A3"
    _autosize_sheet(sheet)


def build_rules_workbook(path=RULES_WORKBOOK_PATH):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    glossary = wb.create_sheet("Glossary")
    _write_glossary(glossary)

    instr = wb.create_sheet("How to Use")
    _write_instructions_sheet(instr)

    comp = wb.create_sheet("01 Salary Components")
    r = 1
    r = _write_table(comp, r, "Basic Salary", ["Monthly CTC"], ["Basic"], [["1", "-", "decimal(Employee.monthly_ctc * 0.4, 0)"]])
    r = _write_table(comp, r, "HRA", ["Basic"], ["HRA"], [["1", "-", "decimal(Salary.basic * 0.4, 0)"]])
    r = _write_table(comp, r, "Fixed Allowances", [], ["Transport Allowance", "Medical Allowance"], [["1", "1600", "1250"]])
    r = _write_table(comp, r, "Bonus", ["Basic"], ["Bonus"], [["1", "-", "decimal(Salary.basic * 0.0833, 0)"]])
    r = _write_table(
        comp,
        r,
        "Employer PF",
        ["PF Enabled", "Basic"],
        ["Employer PF"],
        [["1", '"No"', "-", "0"], ["2", '"Yes"', "-", "decimal(min(Salary.basic, 15000) * 0.12, 0)"]],
    )
    r = _write_table(comp, r, "Gratuity", ["Basic"], ["Gratuity"], [["1", "-", "decimal(Salary.basic * 0.0481, 0)"]])
    r = _write_table(comp, r, "Employer Insurance", [], ["Employer Insurance"], [["1", "1000"]])
    r = _write_table(
        comp,
        r,
        "Employee PF",
        ["PF Enabled", "Basic"],
        ["Employee PF"],
        [["1", '"No"', "-", "0"], ["2", '"Yes"', "-", "decimal(min(Salary.basic, 15000) * 0.12, 0)"]],
    )
    r = _write_table(
        comp,
        r,
        "Gross Before ESI",
        ["Monthly CTC", "Employer PF", "Gratuity", "Employer Insurance"],
        ["Gross Before ESI"],
        [["1", "-", "-", "-", "-", "decimal(Employee.monthly_ctc - EmployerCost.employer_pf - EmployerCost.gratuity - EmployerCost.employer_insurance, 0)"]],
    )
    r = _write_table(
        comp,
        r,
        "Employer ESI",
        ["Gross Before ESI"],
        ["Employer ESI"],
        [["1", "<= 21000", "decimal((Salary.gross_before_esi / 1.0325) * 0.0325, 0)"], ["2", "> 21000", "0"]],
    )
    r = _write_table(comp, r, "Gross", ["Gross Before ESI", "Employer ESI"], ["Gross"], [["1", "-", "-", "decimal(Salary.gross_before_esi - EmployerCost.employer_esi, 0)"]])
    r = _write_table(comp, r, "Employee ESI", ["Gross"], ["Employee ESI"], [["1", "<= 21000", "decimal(Salary.gross * 0.0075, 0)"], ["2", "> 21000", "0"]])
    r = _write_table(
        comp,
        r,
        "Professional Tax",
        ["State", "Gross"],
        ["Professional Tax"],
        [["1", '"Delhi"', "<= 15000", "0"], ["2", '"Delhi"', "> 15000", "200"], ["3", "-", "-", "0"]],
        hit_policy="F",
    )
    _write_table(
        comp,
        r,
        "Special Allowance",
        ["Gross", "Basic", "HRA", "Transport Allowance", "Medical Allowance", "Bonus"],
        ["Special"],
        [["1", "-", "-", "-", "-", "-", "-", "decimal(Salary.gross - Salary.basic - Salary.hra - Salary.transport_allowance - Salary.medical_allowance - Salary.bonus, 0)"]],
    )
    comp.freeze_panes = "A2"
    _autosize_sheet(comp)

    tax = wb.create_sheet("02 Tax and TDS")
    r = 1
    r = _write_table(tax, r, "Taxable Income", ["Gross", "Other Deductions"], ["Taxable Annual"], [["1", "-", "-", "decimal(max(0, (Salary.gross * 12) - 75000 - Employee.other_deductions), 0)"]])
    r = _write_table(
        tax,
        r,
        "Income Tax Slabs",
        ["Taxable Annual"],
        ["Tax Before Rebate"],
        [
            ["1", "<= 400000", "0"],
            ["2", "(400000..800000]", "decimal((Tax.taxable_annual - 400000) * 0.05, 0)"],
            ["3", "(800000..1200000]", "decimal(20000 + (Tax.taxable_annual - 800000) * 0.10, 0)"],
            ["4", "(1200000..1600000]", "decimal(60000 + (Tax.taxable_annual - 1200000) * 0.15, 0)"],
            ["5", "(1600000..2000000]", "decimal(120000 + (Tax.taxable_annual - 1600000) * 0.20, 0)"],
            ["6", "(2000000..2400000]", "decimal(200000 + (Tax.taxable_annual - 2000000) * 0.25, 0)"],
            ["7", "> 2400000", "decimal(300000 + (Tax.taxable_annual - 2400000) * 0.30, 0)"],
        ],
    )
    r = _write_table(tax, r, "Tax Rebate", ["Taxable Annual", "Tax Before Rebate"], ["Tax After Rebate"], [["1", "<= 1200000", "-", "0"], ["2", "> 1200000", "-", "Tax.tax_before_rebate"]])
    _write_table(tax, r, "Final TDS", ["Tax After Rebate"], ["TDS"], [["1", "-", "decimal(decimal(Tax.tax_after_rebate * 1.04, 0) / 12, 0)"]])
    tax.freeze_panes = "A2"
    _autosize_sheet(tax)

    finals = wb.create_sheet("03 Final Salary")
    _write_table(
        finals,
        1,
        "Final Outputs",
        ["Gross", "Employee PF", "Employee ESI", "Professional Tax", "TDS", "Employer PF", "Employer ESI", "Gratuity", "Employer Insurance"],
        ["Take Home", "CTC (Total)"],
        [[
            "1",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "decimal(Salary.gross - Deduction.employee_pf - Deduction.employee_esi - Deduction.professional_tax - Tax.tds, 0)",
            "decimal(Salary.gross + EmployerCost.employer_pf + EmployerCost.employer_esi + EmployerCost.gratuity + EmployerCost.employer_insurance, 0)",
        ]],
    )
    finals.freeze_panes = "A2"
    _autosize_sheet(finals)

    wb.active = wb.sheetnames.index("How to Use")
    wb.save(path)
    return path
