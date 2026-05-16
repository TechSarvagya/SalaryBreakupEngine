from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.config import EMPLOYEE_TEMPLATE_PATH

INPUT_COLUMNS = [
    "Employee ID",
    "Employee Name",
    "Monthly CTC",
    "CCA",
    "PF Enabled",
    "State",
    "Other Deductions",
]


def _write_employee_template() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employees"
    for col, heading in enumerate(INPUT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col, value=heading)
        cell.font = Font(bold=True)
    workbook.save(EMPLOYEE_TEMPLATE_PATH)


def ensure_employee_template() -> None:
    EMPLOYEE_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not EMPLOYEE_TEMPLATE_PATH.exists():
        _write_employee_template()
        return

    workbook = load_workbook(EMPLOYEE_TEMPLATE_PATH, read_only=True)
    sheet = workbook.active
    headings = [sheet.cell(row=1, column=index).value for index in range(1, len(INPUT_COLUMNS) + 1)]
    workbook.close()
    if headings != INPUT_COLUMNS:
        _write_employee_template()


def read_employee_workbook(path: Path) -> pd.DataFrame:
    return pd.read_excel(path)
